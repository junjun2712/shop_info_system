import openpyxl
from entry.commodity import commodity
from Dao.commodityMapper import adds_com

class add_com_excel:
    def __init__(self, path):
        self.filepath = path
        self.list_com = []
        self.result = None
        self.add_data()

    def add_data(self):
        inwb = openpyxl.load_workbook(self.filepath)
        ws = inwb[inwb.sheetnames[0]]
        excel_max_row = ws.max_row

        for row in range(2, excel_max_row + 1):
            com_code = ws.cell(row, 1).value,
            com_name = ws.cell(row, 2).value,
            com_price = ws.cell(row, 3).value,
            com_stock = ws.cell(row, 4).value,
            com = commodity(com_code, com_name, com_price, com_stock)
            com_dict = com.get_commodity()
            self.list_com.append(com_dict)
        result_list = adds_com(self.list_com)
        self.result = result_list[0]